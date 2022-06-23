from csv import excel
import requests
from bs4 import BeautifulSoup
import openpyxl

excel=openpyxl.Workbook()
# print(excel.sheetnames)
sheet=excel.active 
sheet.title='clutch web scraping'
sheet.append(["Company","Website","Location","Rating","Review Count","Hourly Rate","Min Project Size","Employee Size"])
# print(excel.sheetnames)
 
# Making a GET request
url=f"https://clutch.co/directory/mobile-application-developers"
r = requests.get(url,verify=False)

    # Parsing the HTML
soup = BeautifulSoup(r.text, 'html.parser')
pageNumber=soup.find('ul',class_="pagination justify-content-center").find('li', class_="page-item last").a['href']
pageNumber=pageNumber.split('=')[1] 
for i in range(int(pageNumber)+1):
    try:
        url=f"https://clutch.co/directory/mobile-application-developers?page={i}"
        r = requests.get(url,verify=False)
        print(r.status_code) 
        # Parsing the HTML
        soup = BeautifulSoup(r.text, 'html.parser')
        companies=soup.find('ul', class_="directory-list shortlist").find_all('li',class_=["provider provider-row","provider provider-row sponsor"])
        print(url)
        print(len(companies))
        for comp in companies:
            try:
                # print(comp)
                Company=comp.find('div', class_=["company col-md-12 prompt-target","company col-md-12 prompt-target sponsor"]).find('h3', class_="company_info").a.text.strip()
                print(Company)
                emp_size=comp.find('div', class_="module-list").find_all('span')
                # print(emp_size)
                resss=[]
                for xh in emp_size:
                    resss.append(str(xh))
                # print(resss)
                min_proj_size=resss[0].strip('<span>').strip('</')
                Location=resss[3].split('>')[1].split('<')[0]
                hourly_rate=resss[1].strip('<span>').strip('</')
                employees=resss[2].strip('<span>').strip('</')
                try:
                    Rating=comp.find('div', class_="rating-reviews sg-rating").span.text.strip()
                    review_count=comp.find('div', class_="rating-reviews sg-rating").find('a', class_="reviews-link sg-rating__reviews").get_text(strip=True)
                except Exception as noneerror:
                    print(noneerror)
                    rating=''
                    review_count=''
                Website=comp.find('li', class_="website-link website-link-a").a['href']
                sheet.append([Company,Website,Location,Rating,review_count,hourly_rate,min_proj_size,employees])
            except Exception as companyInfoError:
                print(companyInfoError)    
        excel.save('web scarping clutch.co.xlsx')
    except Exception as err:
        print(err)
print("Scrapping Done")